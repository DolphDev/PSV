"""This file contains all csv/excel parsing code"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from .utils import multireplace

def parser(csvfile, cls, *args, **kwargs):
    """This generates row objects for csv, and sets them up 
    for dynamic access"""
    for row in csvfile:
        yield cls({multireplace(x.lower(), " ", "(", ")"): {"org_name": x, "value": row[x]}
        for x in row.keys()}, *args, **kwargs)

def parser_addrow(columns, cls, *args, **kwargs):
    r = cls({}, *args, **kwargs)
    r.update(({multireplace(x.lower(), " ", "(", ")"): {"org_name": x, "value": ""} for x in columns}))
    return r


def xlsxparser_gen(wb2, maxrow, max_column, columnnames, cls):
    """Generator for xlsx parsing"""
    count = 0
    for x in wb2.iter_rows("A2:{}{}".format(max_column, maxrow)):
        d = {}
        for y in columnnames:
            d.update({multireplace(y, " ", "(", ")").lower(): {"org_name": y, "value": x[count].value}})
            count += 1
        yield cls(d)
        count = 0

def excelparser(max_column, max_row):
    def xlsxparser(excelfile, cls, max_column=max_column, max_row=max_row):
        """The Excel Spreadsheet contains many Junk rows/columnnames
        This processes it to be compatible with the program"""
        wb = load_workbook(excelfile)
        wb2 = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        columnnames = []
        for x in (wb2.iter_rows("A1:{}1".format(max_column))):
            for y in x:
                columnnames.append(y.value) 
        gen = xlsxparser_gen(wb2, max_row, max_column, columnnames, cls)
        return gen
    return excelparser
